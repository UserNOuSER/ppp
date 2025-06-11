import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font

class Category:
    def __init__(self, category_id, name):
        self.category_id = category_id
        self.name = name

class Product:
    def __init__(self, product_id, name, category_id, price):
        self.product_id = product_id
        self.name = name
        self.category_id = category_id
        self.price = price

class Position:
    def __init__(self, position_id, name, salary):
        self.position_id = position_id
        self.name = name
        self.salary = salary

class Employee:
    def __init__(self, employee_id, name, position_id):
        self.employee_id = employee_id
        self.name = name
        self.position_id = position_id

class Sale:
    def __init__(self, sale_id, product_id, employee_id, date, quantity):
        self.sale_id = sale_id
        self.product_id = product_id
        self.employee_id = employee_id
        self.date = date
        self.quantity = quantity


class DataGenerator:
    @staticmethod
    def generate_categories():
        categories = ["Смартфоны", "Ноутбуки", "Планшеты", "Наушники", "Аксессуары"]
        return [Category(i+1, name) for i, name in enumerate(categories)]

    @staticmethod
    def generate_products(categories):
        products = []
        names = [
            "Galaxy S", "iPhone", "MacBook", "ThinkPad", "iPad",
            "AirPods", "Watch", "Зарядка", "Чехол", "Кабель"
        ]
        for i in range(50):
            product_id = i + 1
            name = f"{random.choice(names)} {random.randint(1, 20)}"
            category_id = random.choice(categories).category_id
            price = random.randint(5000, 150000)
            products.append(Product(product_id, name, category_id, price))
        return products

    @staticmethod
    def generate_positions():
        positions = ["Менеджер", "Продавец", "Консультант", "Администратор"]
        return [Position(i+1, name, random.randint(30000, 80000)) for i, name in enumerate(positions)]

    @staticmethod
    def generate_employees(positions):
        employees = []
        names = ["Иван", "Мария", "Алексей", "Ольга", "Дмитрий", "Анна"]
        surnames = ["Иванов", "Петрова", "Сидоров", "Кузнецова", "Смирнов", "Васильева"]
        for i in range(15):
            employee_id = i + 1
            name = f"{random.choice(names)} {random.choice(surnames)}"
            position_id = random.choice(positions).position_id
            employees.append(Employee(employee_id, name, position_id))
        return employees

    @staticmethod
    def generate_sales(products, employees):
        sales = []
        start_date = datetime(2023, 1, 1)
        for i in range(200):
            sale_id = i + 1
            product = random.choice(products)
            employee = random.choice(employees)
            date = start_date + timedelta(days=random.randint(0, 365))
            quantity = random.randint(1, 5)
            sales.append(Sale(sale_id, product.product_id, employee.employee_id, date, quantity))
        return sales


class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def add_sheet(self, name, data, headers):
        ws = self.wb.create_sheet(name)
        ws.append(headers)
        for item in data:
            ws.append([getattr(item, attr) for attr in headers])

        # Форматирование заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)

    def save(self):
        self.wb.save(self.filename)
        print(f"Файл {self.filename} успешно создан!")

    def load_data(self):
        wb = openpyxl.load_workbook(self.filename)
        data = {}

        # Загрузка категорий
        ws = wb['Categories']
        data['categories'] = [Category(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True)]

        # Загрузка продуктов
        ws = wb['Products']
        data['products'] = [Product(row[0], row[1], row[2], row[3]) for row in
                            ws.iter_rows(min_row=2, values_only=True)]

        # Загрузка должностей
        ws = wb['Positions']
        data['positions'] = [Position(row[0], row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)]

        # Загрузка сотрудников
        ws = wb['Employees']
        data['employees'] = [Employee(row[0], row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)]

        # Загрузка продаж
        ws = wb['Sales']
        data['sales'] = [Sale(row[0], row[1], row[2], row[3], row[4]) for row in
                         ws.iter_rows(min_row=2, values_only=True)]

        return data


class DataAnalyzer:
    @staticmethod
    def sales_report(data):
        """Продажи с детализацией по товарам и сотрудникам"""
        report = []
        for sale in data['sales']:
            product = next(p for p in data['products'] if p.product_id == sale.product_id)
            category = next(c for c in data['categories'] if c.category_id == product.category_id)
            employee = next(e for e in data['employees'] if e.employee_id == sale.employee_id)
            position = next(p for p in data['positions'] if p.position_id == employee.position_id)

            report.append({
                'Дата': sale.date.strftime("%d.%m.%Y"),
                'Товар': product.name,
                'Категория': category.name,
                'Цена': product.price,
                'Кол-во': sale.quantity,
                'Сотрудник': employee.name,
                'Должность': position.name,
                'Сумма': product.price * sale.quantity
            })
        return report

    @staticmethod
    def category_stats(data):
        """Статистика продаж по категориям"""
        stats = {}
        for sale in data['sales']:
            product = next(p for p in data['products'] if p.product_id == sale.product_id)
            category = next(c for c in data['categories'] if c.category_id == product.category_id)

            if category.name not in stats:
                stats[category.name] = {'total_sales': 0, 'total_revenue': 0}

            stats[category.name]['total_sales'] += sale.quantity
            stats[category.name]['total_revenue'] += sale.quantity * product.price
        return stats

    @staticmethod
    def employee_performance(data):
        """Эффективность сотрудников по должностям"""
        performance = {}
        for employee in data['employees']:
            position = next(p for p in data['positions'] if p.position_id == employee.position_id)

            if position.name not in performance:
                performance[position.name] = {'employees': 0, 'total_sales': 0, 'total_revenue': 0}

            performance[position.name]['employees'] += 1

            for sale in data['sales']:
                if sale.employee_id == employee.employee_id:
                    product = next(p for p in data['products'] if p.product_id == sale.product_id)
                    performance[position.name]['total_sales'] += sale.quantity
                    performance[position.name]['total_revenue'] += sale.quantity * product.price

        # Расчет средних показателей
        for pos in performance:
            if performance[pos]['employees'] > 0:
                performance[pos]['avg_sales'] = performance[pos]['total_sales'] / performance[pos]['employees']
                performance[pos]['avg_revenue'] = performance[pos]['total_revenue'] / performance[pos]['employees']
        return performance


def main():
    # Генерация данных
    categories = DataGenerator.generate_categories()
    products = DataGenerator.generate_products(categories)
    positions = DataGenerator.generate_positions()
    employees = DataGenerator.generate_employees(positions)
    sales = DataGenerator.generate_sales(products, employees)

    # Создание Excel файла
    excel = ExcelManager("electronics_sales.xlsx")
    excel.add_sheet("Categories", categories, ["category_id", "name"])
    excel.add_sheet("Products", products, ["product_id", "name", "category_id", "price"])
    excel.add_sheet("Positions", positions, ["position_id", "name", "salary"])
    excel.add_sheet("Employees", employees, ["employee_id", "name", "position_id"])
    excel.add_sheet("Sales", sales, ["sale_id", "product_id", "employee_id", "date", "quantity"])
    excel.save()

    # Загрузка данных для анализа
    loaded_data = excel.load_data()

    # Выполнение запросов
    analyzer = DataAnalyzer()

    print("\nОтчет о продажах (первые 5 записей):")
    sales_report = analyzer.sales_report(loaded_data)
    for entry in sales_report[:5]:
        print(entry)

    print("\nСтатистика по категориям:")
    category_stats = analyzer.category_stats(loaded_data)
    for category, stats in category_stats.items():
        print(f"{category}: {stats['total_sales']} прод., {stats['total_revenue']} руб.")

    print("\nЭффективность сотрудников:")
    performance = analyzer.employee_performance(loaded_data)
    for position, stats in performance.items():
        print(f"{position}:")
        print(f"  Средние продажи: {stats['avg_sales']:.1f} шт./сотр.")
        print(f"  Средняя выручка: {stats['avg_revenue']:.1f} руб./сотр.")


if __name__ == "__main__":
    main()
    